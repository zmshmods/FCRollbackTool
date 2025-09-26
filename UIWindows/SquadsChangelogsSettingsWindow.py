import sys
import json
import os
import csv
import openpyxl
import io 
from typing import Optional

from PySide6.QtWidgets import (
    QApplication, QVBoxLayout, QHBoxLayout, QLabel, QWidget, QSizePolicy, QPushButton, QFileDialog
)
from PySide6.QtCore import Qt
from PySide6.QtGui import QGuiApplication, QIcon
from qfluentwidgets import CheckBox, ComboBox, LineEdit, Theme, setTheme, setThemeColor, SimpleCardWidget#, FluentIcon, CaptionLabel

from UIComponents.Tooltips import apply_tooltip
from UIComponents.Personalization import BaseWindow
from UIComponents.TitleBar import TitleBar
from UIComponents.MainStyles import MainStyles

from Core.Logger import logger
from Core.ConfigManager import ConfigManager
from Core.GameManager import GameManager
from Core.ErrorHandler import ErrorHandler

# Constants
WINDOW_TITLE = "Changelogs Settings"
WINDOW_SIZE = (520, 250)
THEME_COLOR = "#00FF00"
ICON_PATH = "Data/Assets/Icons/FRICON.png"
SPACER_WIDTH = 75
BAR_HEIGHT = 32
SHOW_MAX_BUTTON = False
SHOW_MIN_BUTTON = False
SHOW_CLOSE_BUTTON = False

# Unified styles
TITLE_STYLE = "font-weight: bold; color: white; font-size: 16px;"
DESC_STYLE = "color: rgba(255, 255, 255, 0.7); font-size: 12px;"
TEXT_STYLE = "background-color: transparent; color: white; font-size: 14px;"
SEPARATOR_STYLE = "background-color: rgba(255, 255, 255, 0.1);"

# Settings options
SAVE_TYPE_ITEMS = [".xlsx", ".csv", ".json"]
FORMAT_CONFIG = {
    ".xlsx": (".xlsx", lambda c, p: c.to_xlsx(p)),
    ".csv": (".csv", lambda c, p: c.to_csv(p)),
    ".json": (".json", lambda c, p: c.to_json(p))
}

class ChangelogsSettings:
    """Handles changelog data processing and formatting based on user settings."""
    def __init__(self, xlsx_data: bytes, changelog_name: str, config_manager: ConfigManager,
                 changelog_info: dict = None, index_url: str = None):
        self.xlsx_data = xlsx_data
        self.changelog_name = changelog_name
        self.config_manager = config_manager
        self.changelog_info = changelog_info or {}
        self.index_url = index_url
        self.game_manager = GameManager()

    def to_xlsx(self, output_path: str) -> str:
        """Save the changelog data as XLSX without modification."""
        try:
            os.makedirs(os.path.dirname(output_path), exist_ok=True)
            with open(output_path, "wb") as f:
                f.write(self.xlsx_data)
            return output_path
        except Exception as e:
            raise Exception(f"Failed to save XLSX file: {str(e)}")

    def to_csv(self, output_path: str) -> str:
        """Convert Excel sheets to separate CSV files in a folder named after the changelog type."""
        try:
            changelog_type = self.changelog_info.get("Type", "UnknownType")
            base_dir = os.path.join(os.path.dirname(output_path), changelog_type)
            os.makedirs(base_dir, exist_ok=True)
            # Use BytesIO to load raw bytes data
            workbook = openpyxl.load_workbook(io.BytesIO(self.xlsx_data))
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheet_path = os.path.join(base_dir, f"{sheet_name}.csv")
                with open(sheet_path, "w", encoding='utf-8-sig', newline='') as f:
                    writer = csv.writer(f)
                    for row in sheet.iter_rows(values_only=True):
                        writer.writerow(row)
            return base_dir
        except Exception as e:
            raise Exception(f"Failed to convert to CSV: {str(e)}")

    def to_json(self, output_path: str) -> str:
        """Convert Excel sheets to separate JSON files in a folder named after the changelog type."""
        try:
            changelog_type = self.changelog_info.get("Type", "UnknownType")
            base_dir = os.path.join(os.path.dirname(output_path), changelog_type)
            os.makedirs(base_dir, exist_ok=True)
            workbook = openpyxl.load_workbook(io.BytesIO(self.xlsx_data))
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                json_path = os.path.join(base_dir, f"{sheet_name}.json")

                rows = list(sheet.iter_rows(values_only=True))
                if not rows:
                    continue
                headers = rows[0]
                data = []
                for row in rows[1:]:
                    entry = dict(zip(headers, row))
                    data.append(entry)

                with open(json_path, "w", encoding='utf-8') as f:
                    json.dump(data, f, ensure_ascii=False, indent=2)
            return base_dir
        except Exception as e:
            raise Exception(f"Failed to convert to JSON: {str(e)}")

class ChangelogsSettingsWindow(BaseWindow):
    def __init__(self, parent: Optional['QWidget'] = None):
        super().__init__(parent=parent)
        self.config_manager = ConfigManager()
        self.button_manager = ButtonManager(self)
        self.save_type_combo = None 
        self.setWindowTitle(WINDOW_TITLE)
        self.resize(*WINDOW_SIZE)
        self.setWindowModality(Qt.ApplicationModal)
        self.center_window()
        self.main_layout = QVBoxLayout(self)
        self.main_layout.setContentsMargins(0, 0, 0, 0)
        self.main_layout.setSpacing(0)
        self._setup_ui()
        try:
            self.button_manager._update_reset_button_state()
        except Exception as e:
            ErrorHandler.handleError(f"Failed to setup UI: {str(e)}")

    def _setup_ui(self) -> None:
        """Initialize the UI components."""
        self._setup_title_bar()
        self._setup_content_container()
        self._setup_buttons()
        
    def center_window(self) -> None:
        """Center the window on the screen."""
        screen = QGuiApplication.primaryScreen().geometry()
        self.move((screen.width() - self.width()) // 2, (screen.height() - self.height()) // 2)

    def _setup_title_bar(self) -> None:
        """Configure the title bar."""
        title_bar = TitleBar(
            window=self,
            title=WINDOW_TITLE,
            icon_path=ICON_PATH,
            spacer_width=SPACER_WIDTH,
            show_max_button=SHOW_MAX_BUTTON,
            show_min_button=SHOW_MIN_BUTTON,
            show_close_button=SHOW_CLOSE_BUTTON,
            bar_height=BAR_HEIGHT
        )
        title_bar.create_title_bar()

    def _setup_content_container(self) -> None:
        """Setup the main content container with settings widgets."""
        self.content_container = QWidget()
        self.content_layout = QVBoxLayout(self.content_container)
        self.content_layout.setContentsMargins(10, 10, 10, 10)
        self.content_layout.setSpacing(0)

        #desc = CaptionLabel("Configure Changelogs Settings", self)
        #desc.setStyleSheet(DESC_STYLE)
        #self.content_layout.addWidget(desc)

        #self.content_layout.addSpacerItem(QSpacerItem(0, 10, QSizePolicy.Minimum, QSizePolicy.Fixed))

        content_widget = self._create_changelogs_settings_content()
        self.content_layout.addWidget(content_widget)

        self.content_layout.addStretch()
        self.main_layout.addWidget(self.content_container)

    def _setup_buttons(self) -> None:
        """Setup the buttons container."""
        separator = self._create_separator(height=1)
        self.main_layout.addWidget(separator)
        button_container = self.button_manager.create_buttons()
        self.main_layout.addWidget(button_container if button_container else QWidget())

    def _create_separator(self, width: Optional[int] = None, height: Optional[int] = None) -> QWidget:
        """Create a separator widget with specified dimensions."""
        separator = QWidget()
        separator.setStyleSheet(SEPARATOR_STYLE)
        if width is not None:
            separator.setFixedWidth(width)
        if height is not None:
            separator.setFixedHeight(height)
        return separator

    def _create_changelogs_settings_content(self) -> QWidget:
        """Create the content for changelog settings."""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(10)

        card = SimpleCardWidget()
        card_layout = QVBoxLayout(card)
        card_layout.setContentsMargins(10, 10, 10, 10)
        card_layout.setSpacing(10)

        # Save As Type
        save_type_container = QWidget()
        save_type_layout = QHBoxLayout(save_type_container)
        save_type_layout.setContentsMargins(0, 0, 0, 0)
        save_type_layout.setSpacing(0)

        save_type_label = QLabel("Save As Type:")
        save_type_label.setStyleSheet(TEXT_STYLE)

        self.save_type_combo = ComboBox()
        self.save_type_combo.addItems(SAVE_TYPE_ITEMS)
        self.save_type_combo.setCurrentText(self.config_manager.getConfigKeyChangelogFormat())
        self.save_type_combo.setFixedSize(300, 28)
        apply_tooltip(self.save_type_combo, "saveTypeOptions")
        self.save_type_combo.activated.connect(lambda index: self.on_save_type_changed(self.save_type_combo.itemText(index)))

        save_type_layout.addWidget(save_type_label)
        save_type_layout.addStretch()
        save_type_layout.addWidget(self.save_type_combo)

        card_layout.addWidget(save_type_container)

        # Separator
        separator = self._create_separator(height=1)
        card_layout.addWidget(separator)

        # Save Path
        save_path_container = QWidget()
        save_path_layout = QVBoxLayout(save_path_container)
        save_path_layout.setContentsMargins(0, 0, 0, 0)
        save_path_layout.setSpacing(10)

        # Horizontal row for label and button
        top_row = QWidget()
        top_layout = QHBoxLayout(top_row)
        top_layout.setContentsMargins(0, 0, 0, 0)
        save_path_label = QLabel("Save Path:")
        save_path_label.setStyleSheet(TEXT_STYLE)

        self.browse_button = QPushButton("Browse")
        self.browse_button.setFixedSize(300, 28)
        apply_tooltip(self.browse_button, "browseButton")
        self.browse_button.clicked.connect(self.on_browse)

        top_layout.addWidget(save_path_label)
        top_layout.addStretch()
        top_layout.addWidget(self.browse_button)

        # Text field below
        self.save_path_edit = LineEdit()
        self.save_path_edit.setMinimumHeight(30)
        self.save_path_edit.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.save_path_edit.setPlaceholderText("Set path or leave it empty to be asked each time")
        self.save_path_edit.setText(self.config_manager.getConfigKeyChangelogSavePath() or "")
        self.save_path_edit.textChanged.connect(self.on_save_path_changed)

        save_path_layout.addWidget(top_row)
        save_path_layout.addWidget(self.save_path_edit)

        card_layout.addWidget(save_path_container)

        # Separator
        separator = self._create_separator(height=1)
        card_layout.addWidget(separator)

        # Squad Folder Checkbox
        self.squad_folder_check = CheckBox("Save changelogs in a subfolder named using the squad file name")
        self.squad_folder_check.setStyleSheet(TEXT_STYLE)
        self.squad_folder_check.setChecked(self.config_manager.getConfigKeySaveChangelogsInFolderUsingSquadFileName())
        self.squad_folder_check.stateChanged.connect(self.on_squad_folder_changed)

        card_layout.addWidget(self.squad_folder_check)

        layout.addWidget(card)
        layout.addStretch()
        return widget

    def on_save_type_changed(self, format_type: str) -> None:
        """Handle save type selection changes."""
        try:
            if format_type not in FORMAT_CONFIG:
                format_type = ".xlsx"  # Default fallback
                logger.warning(f"Invalid save type '{format_type}', falling back to .xlsx")
            self.config_manager.setConfigKeyChangelogFormat(format_type)
            self.button_manager._update_reset_button_state()
        except Exception as e:
            ErrorHandler.handleError(f"Failed to update changelog format: {str(e)}")

    def on_save_path_changed(self, save_path: str) -> None:
        """Handle save path changes."""
        try:
            self.config_manager.setConfigKeyChangelogSavePath(save_path)
            self.button_manager._update_reset_button_state()
        except Exception as e:
            ErrorHandler.handleError(f"Failed to update save path: {str(e)}")

    """def _is_valid_path(self, path: str) -> bool:
        logger.info(f"Validating path: '{path}'")
        try:
            p = Path(path)
            is_valid = p.is_absolute() and p.exists()
            logger.info(f"Path is absolute? {'Yes' if p.is_absolute() else 'No'}")
            logger.info(f"Path exists? {'Yes' if p.exists() else 'No'}")
            return is_valid
        except Exception as e:
            ErrorHandler.handleError(f"Exception in _is_valid_path: {str(e)}")
            return False"""

    def on_squad_folder_changed(self, state: int) -> None:
        """Handle squad folder checkbox state changes."""
        try:
            checked = bool(state == Qt.CheckState.Checked.value)
            self.config_manager.setConfigKeySaveChangelogsInFolderUsingSquadFileName(checked)
            self.button_manager._update_reset_button_state()
        except Exception as e:
            ErrorHandler.handleError(f"Failed to update changelog folder setting: {str(e)}")

    def on_browse(self) -> None:
        """Handle browse button click for save path."""
        try:
            folder_path = QFileDialog.getExistingDirectory(
                parent=self,
                caption="Select Save Path",
                dir=os.path.join(os.path.expanduser("~"), "Desktop")
            )
            if folder_path:
                self.config_manager.setConfigKeyChangelogSavePath(folder_path)
                self.save_path_edit.setText(folder_path)
                self.button_manager._update_reset_button_state()
        except Exception as e:
            ErrorHandler.handleError(f"Failed to browse save path: {str(e)}")

    def close(self) -> None:
        """Close the window."""
        try:
            QApplication.processEvents()
            super().close()
        except Exception as e:
            ErrorHandler.handleError(f"Failed to close window: {str(e)}")

class ButtonManager:
    """Manages button actions for ChangelogsSettingsWindow."""
    def __init__(self, window: "ChangelogsSettingsWindow"):
        self.window = window
        self.button_container = None
        self.buttons = {}
        self.button_layout = None

    def create_buttons(self) -> Optional[QWidget]:
        """Create and configure buttons."""
        try:
            self._init_buttons()
            self._setup_layout()
            self.button_container = QWidget(self.window)
            self.button_container.setLayout(self.button_layout)
            return self.button_container
        except Exception as e:
            ErrorHandler.handleError(f"Error creating buttons: {str(e)}")
            fallback_container = QWidget(self.window)
            fallback_layout = QHBoxLayout(fallback_container)
            fallback_layout.addStretch()
            logger.warning("Returning fallback button container due to error.")
            return fallback_container

    def _init_buttons(self) -> None:
        """Initialize buttons with consistent styling."""
        button_configs = {
            "reset": "Reset All",
            "done": "Done",
        }
        for name, text in button_configs.items():
            try:
                btn = QPushButton(text)
                if name == "reset":
                    apply_tooltip(btn, "resetAllChangelogsSettings")
                self.buttons[name] = btn
            except Exception as e:
                self.buttons[name] = QPushButton(f"{text} (Error)")

    def _setup_layout(self) -> None:
        """Setup the button layout."""
        self.button_layout = QHBoxLayout()
        self.button_layout.addStretch()
        for name in ["reset", "done"]:
            if name in self.buttons:
                btn = self.buttons[name]
                if name == "reset":
                    btn.clicked.connect(self._reset_to_default)
                elif name == "done":
                    btn.clicked.connect(self.window.close)
                self.button_layout.addWidget(btn)

    def _reset_to_default(self) -> None:
        """Reset all settings to their default values."""
        try:
            # Reset settings in ConfigManager
            self.window.config_manager.resetChangelogSettings()

            # Define UI elements and their update methods
            ui_elements = [
                (self.window.save_type_combo, lambda: self.window.config_manager.getConfigKeyChangelogFormat()),
                (self.window.save_path_edit, lambda: self.window.config_manager.getConfigKeyChangelogSavePath() or ""),
                (self.window.squad_folder_check, lambda: self.window.config_manager.getConfigKeySaveChangelogsInFolderUsingSquadFileName()),
            ]

            # Block signals, update UI, and re-enable signals
            for element, get_value in ui_elements:
                element.blockSignals(True)
                if isinstance(element, ComboBox):
                    element.setCurrentText(get_value())
                elif isinstance(element, LineEdit):
                    element.setText(get_value())
                elif isinstance(element, CheckBox):
                    element.setChecked(get_value())
                element.blockSignals(False)

            # Update reset button state after resetting
            self._update_reset_button_state()
            logger.info("Changelog settings reset to default values.")
        except Exception as e:
            ErrorHandler.handleError(f"Failed to reset changelog settings: {str(e)}")

    def _update_reset_button_state(self) -> None:
        """Enable or disable the Reset button based on whether settings differ from defaults."""
        try:
            default_settings = self.window.config_manager.getDefaultChangelogSettings()
            current_settings = {
                "ChangelogFormat": self.window.config_manager.getConfigKeyChangelogFormat(),
                "ChangelogSavePath": self.window.config_manager.getConfigKeyChangelogSavePath() or "",
                "SaveChangelogsInFolderUsingSquadFileName": self.window.config_manager.getConfigKeySaveChangelogsInFolderUsingSquadFileName(),
                "SelectAllChangelogs": self.window.config_manager.getConfigKeySelectAllChangelogs()
            }
            settings_changed = any(
                current_settings[key] != default_settings.get(key)
                for key in default_settings
            )

            if "reset" in self.buttons:
                self.buttons["reset"].setEnabled(settings_changed)
        except Exception as e:
            ErrorHandler.handleError(f"Failed to update reset button state: {str(e)}")
            if "reset" in self.buttons:
                self.buttons["reset"].setEnabled(True)

def main() -> int:
    try:
        app = QApplication(sys.argv)
        app.setStyleSheet(MainStyles())
        app.setWindowIcon(QIcon(ICON_PATH))
        setTheme(Theme.DARK)
        setThemeColor(THEME_COLOR)
        window = ChangelogsSettingsWindow()
        window.show()
        return app.exec()
    except Exception as e:
        ErrorHandler.handleError(f"Error in main: {str(e)}")
        return 1

if __name__ == "__main__":
    main()